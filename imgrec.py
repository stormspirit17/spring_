from PIL import Image
import numpy as np
from statistics import mode
import matplotlib.pyplot as plt
# i = Image.open('images/numbers/y0.5.png')
# iar = np.asarray(i)
# print(iar)
# plt.imshow(iar)
# plt.show()
def treshold(imageArray):
    balancear = []
    newar = imageArray
    from statistics import mean
    for eachrow in imageArray:
        for eachpix in eachrow:
            avg = mean(eachpix[:3])
            balancear.append(avg)
    balance = mean(balancear)
    for eachrow in newar:
        for eachpix in eachrow:
            if mean(eachpix[:3]) > balance:
                eachpix[0] = 255
                eachpix[1] = 255
                eachpix[2] = 255
                # eachpix[3] = 255
            else:
                eachpix[0] = 0
                eachpix[1] = 0
                eachpix[2] = 0
                # eachpix[3] = 255
    return newar

i = Image.open('v1/frame3.jpg')
i1 = Image.open('v1/frame1699.jpg')
i2 = Image.open('v1/frame1566.jpg')
i3 = Image.open('v1/frame1555.jpg')
i4 = Image.open('v1/frame60.jpg')
iar = np.array(i)
qar = np.array(i1)
qar1 = np.array(i2)
qar2 = np.array(i3)
qar3 = np.array(i4)
qar4 = np.array(Image.open('v1/frame68.jpg'))
qar5 = np.array(Image.open('v1/frame22.jpg'))
qar6 = np.array(Image.open('v1/frame199.jpg'))
qar7 = np.array(Image.open('v1/frame719.jpg'))
qar8 = np.array(Image.open('v1/frame677.jpg'))
# print(iar[307][112])

# print(one)
class Number(object):
    def __init__(self):
        self.zero = qar[277:295, 136:149]  #0
        self.zeror = qar3[247:265, 404:417] #0
        self.zerod = qar[277:295, 136:149] #0
        self.one = iar[204:222, 113:126]  #1
        self.oner = iar[247:265, 404:417]  #1
        self.oned = qar7[277:295, 136:149]  #1
        self.twor = iar[247:265, 376:389]  #2
        self.tworr = qar4[247:265, 418:431]  #2
        self.twod = qar1[277:295, 136:149]  #2
        self.threer = qar1[247:265, 418:431]  #3
        self.threrr = qar6[247:265, 418:431]  #3
        self.four = iar[278:296, 121:134]  #4
        self.fourr = qar7[247:265, 418:431]  #4
        self.five = iar[204:222, 126:139]  #5
        self.fived = iar[278:296, 107:120]  #5
        self.fivedd = qar7[278:296, 121:134]  #5
        self.fiver = iar[247:265, 418:431]  #5
        self.six = iar[204:222, 140:153]  #6
        self.sixr = qar8[247:265, 418:431]  #6
        self.sevenr = qar2[247:265, 418:431]  #7
        self.eightr = qar[247:265, 390:403]  #8
        self.eightrr = qar4[247:265, 390:403]
        self.eightrrr = qar4[247:265, 418:431]  #8
        self.eightd = qar3[277:295, 136:149] #8
        self.niner = qar[247:265, 404:417]  #9
        self.nined = iar[277:295, 136:149] #9
        self.ninerr = iar[247:265, 390:403] #9
    def touplist(self):
        i = 0
        uplist = [[] for _ in range(100)]
        while i < 99:
            up1 = np.array(Image.open('v2/frame%d.jpg' % i))[204:222, 113:126]
            up2 = np.array(Image.open('v2/frame%d.jpg' % i))[204:222, 126:139]
            up3 = np.array(Image.open('v2/frame%d.jpg' % i))[204:222, 140:153]
            i += 1
            if np.allclose(self.one, up1, atol=50) or np.allclose(self.oner, up1, atol=50) == True:
                uplist[i].append(1)
            elif np.allclose(self.four, up1, atol=50) == True:
                uplist[i].append(4)
            elif np.allclose(self.five, up1, atol=50) or np.allclose(self.fived, up1, atol=50) == True:
                uplist[i].append(5)
            elif np.allclose(self.six, up1, atol=50) == True:
                uplist[i].append(6)
            elif np.allclose(self.zero, up1, atol=50) == True:
                uplist[i].append(0)
            elif np.allclose(self.niner, up1, atol=50) or np.allclose(self.nined, up1, atol=50) == True:
                uplist[i].append(9)
            elif np.allclose(self.threer, up1, atol=50) == True:
                uplist[i].append(3)
            elif np.allclose(self.twor, up1, atol=50) == True:
                uplist[i].append(2)
            elif np.allclose(self.eightr, up1, atol=50) == True:
                uplist[i].append(8)
            elif np.allclose(self.sevenr, up1, atol=50) == True:
                uplist[i].append(7)
            if np.allclose(self.one, up2, atol=50) or np.allclose(self.oner, up2, atol=50) == True:
                uplist[i].append(1)
            elif np.allclose(self.four, up2, atol=60) == True:
                uplist[i].append(4)
            elif np.allclose(self.five, up2, atol=60) or np.allclose(self.fived, up2, atol=60) == True:
                uplist[i].append(5)
            elif np.allclose(self.six, up2, atol=60) == True:
                uplist[i].append(6)
            elif np.allclose(self.zero, up2, atol=60) == True:
                uplist[i].append(0)
            elif np.allclose(self.niner, up2, atol=60) or np.allclose(self.nined, up2, atol=60) == True:
                uplist[i].append(9)
            elif np.allclose(self.threer, up2, atol=60) == True:
                uplist[i].append(3)
            elif np.allclose(self.twor, up2, atol=60) == True:
                uplist[i].append(2)
            elif np.allclose(self.eightr, up2, atol=60) == True:
                uplist[i].append(8)
            elif np.allclose(self.sevenr, up2, atol=60) == True:
                uplist[i].append(7)
            if np.allclose(self.one, up3, atol=60) or np.allclose(self.oner, up3, atol=60) == True:
                uplist[i].append(1)
            elif np.allclose(self.four, up3, atol=60) == True:
                uplist[i].append(4)
            elif np.allclose(self.five, up3, atol=60) or np.allclose(self.fived, up3, atol=60) == True:
                uplist[i].append(5)
            elif np.allclose(self.six, up3, atol=60) == True:
                uplist[i].append(6)
            elif np.allclose(self.zero, up3, atol=60) == True:
                uplist[i].append(0)
            elif np.allclose(self.niner, up3, atol=60) or np.allclose(self.nined, up3, atol=60)== True:
                uplist[i].append(9)
            elif np.allclose(self.threer, up3, atol=60) == True:
                uplist[i].append(3)
            elif np.allclose(self.twor, up3, atol=60) == True:
                uplist[i].append(2)
            elif np.allclose(self.eightr, up3, atol=60) == True:
                uplist[i].append(8)
            elif np.allclose(self.sevenr, up3, atol=60) == True:
                uplist[i].append(7)
        uplist.remove([])
        uplistik = []
        for i in uplist:
            magic = lambda nums: sum(digit * 10 ** (len(nums) - 1 - j)
            for j, digit in enumerate(i))
            uplistik.append(magic(i))
        # count = 0
        # kek = 0
        # for n in uplistik:
        #     if abs(mode(uplistik) - n) > 30:
        #         count += 1
        # uplistik = [x for x in uplistik if abs(mode(uplistik) - x) < 30]
        # print(count)
        # while kek < count +1:
        #     uplistik.append(int(np.mean(uplistik)))
        #     kek += 1
        # print('u', len(uplistik))
        return uplistik
    def todownlist(self):
        i = 0
        downlist = [[] for _ in range(1680)]
        while i < 1679:
            down1 = np.array(Image.open('v2/frame%d.jpg' % i))[278:296, 107:120]
            down2 = np.array(Image.open('v2/frame%d.jpg' % i))[278:296, 121:134]
            down3 = np.array(Image.open('v2/frame%d.jpg' % i))[277:295, 136:149]
            i += 1
            if np.allclose(self.one, down1, atol=27) or np.allclose(self.oner, down1, atol=27) or np.allclose(self.oned, down1, atol=27) == True:
                downlist[i].append(1)
            elif np.allclose(self.four, down1, atol=33) == True:
                downlist[i].append(4)
            elif np.allclose(self.five, down1, atol=38) or np.allclose(self.fived, down1, atol=38) or np.allclose(self.fivedd, down1, atol=38) == True:
                downlist[i].append(5)
            elif np.allclose(self.six, down1, atol=33) == True:
                downlist[i].append(6)
            elif np.allclose(self.zero, down1, atol=33) or np.allclose(self.zerod, down1, atol=33) == True:
                downlist[i].append(0)
            elif np.allclose(self.niner, down1, atol=33) or np.allclose(self.nined, down1, atol=33) == True:
                downlist[i].append(9)
            elif np.allclose(self.threer, down1, atol=33) == True:
                downlist[i].append(3)
            elif np.allclose(self.twor, down1, atol=33) or np.allclose(self.twod, down1, atol=33) == True:
                downlist[i].append(2)
            elif np.allclose(self.eightr, down1, atol=33) == True:
                downlist[i].append(8)
            elif np.allclose(self.sevenr, down1, atol=33) == True:
                downlist[i].append(7)
            if np.allclose(self.one, down2, atol=33) or np.allclose(self.oner, down2, atol=33) or np.allclose(self.oned, down2, atol=33) == True:
                downlist[i].append(1)
            elif np.allclose(self.four, down2, atol=33) == True:
                downlist[i].append(4)
            elif np.allclose(self.five, down2, atol=33) or np.allclose(self.fived, down2, atol=33) or np.allclose(self.fivedd, down2, atol=38) == True:
                downlist[i].append(5)
            elif np.allclose(self.six, down2, atol=33) == True:
                downlist[i].append(6)
            elif np.allclose(self.zero, down2, atol=33) or np.allclose(self.zerod, down2, atol=33) == True:
                downlist[i].append(0)
            elif np.allclose(self.niner, down2, atol=33) or np.allclose(self.nined, down2, atol=33) == True:
                downlist[i].append(9)
            elif np.allclose(self.threer, down2, atol=33) == True:
                downlist[i].append(3)
            elif np.allclose(self.twor, down2, atol=33) or np.allclose(self.twod, down2, atol=33) == True:
                downlist[i].append(2)
            elif np.allclose(self.eightr, down2, atol=33) == True:
                downlist[i].append(8)
            elif np.allclose(self.sevenr, down2, atol=33) == True:
                downlist[i].append(7)
            if np.allclose(self.one, down3, atol=40) or np.allclose(self.oner, down3, atol=40) or np.allclose(self.oned, down3, atol=40) == True:
                downlist[i].append(1)
            elif np.allclose(self.four, down3, atol=40) == True:
                downlist[i].append(4)
            elif np.allclose(self.five, down3, atol=40) or np.allclose(self.fived, down3, atol=40) or np.allclose(self.fivedd, down3, atol=40) == True:
                downlist[i].append(5)
            elif np.allclose(self.six, down3, atol=40) == True:
                downlist[i].append(6)
            elif np.allclose(self.zero, down3, atol=40) or np.allclose(self.zerod, down2, atol=40) == True:
                downlist[i].append(0)
            elif np.allclose(self.niner, down3, atol=40) or np.allclose(self.nined, down3, atol=40)== True:
                downlist[i].append(9)
            elif np.allclose(self.threer, down3, atol=40) == True:
                downlist[i].append(3)
            elif np.allclose(self.twor, down3, atol=40) or np.allclose(self.twod, down3, atol=40) == True:
                downlist[i].append(2)
            elif np.allclose(self.eightr, down3, atol=40) or np.allclose(self.eightd, down3, atol=40) == True:
                downlist[i].append(8)
            elif np.allclose(self.sevenr, down3, atol=40) == True:
                downlist[i].append(7)
        downlist.remove([])
        downlistik = []
        for i in downlist:
            magic = lambda nums: sum(digit * 10 ** (len(nums) - 1 - j)
            for j, digit in enumerate(i))
            downlistik.append(magic(i))
        print(np.mean(downlistik))
        count = 0
        kek = 0
        for n in downlistik:
            if abs(mode(downlistik) - n) > 30:
                count += 1
        downlistik = [x for x in downlistik if abs(mode(downlistik) - x) < 30]
        print(count)
        while kek < count:
            downlistik.append(int(np.mean(downlistik)))
            kek += 1
        print('d', len(downlistik))
        return downlistik
    def torightlist(self):
        i = 0
        rlist = [[] for _ in range(1680)]
        while i < 1679:
            right1 = np.array(Image.open('v2/frame%d.jpg' % i))[247:265, 376:389]
            right2 = np.array(Image.open('v2/frame%d.jpg' % i))[247:265, 390:403]
            right3 = np.array(Image.open('v2/frame%d.jpg' % i))[247:265, 404:417]
            right4 = np.array(Image.open('v2/frame%d.jpg' % i))[247:265, 418:431]
            i += 1
            if np.allclose(self.one, right1, atol=40) or np.allclose(self.oner, right1, atol=40) == True:
                rlist[i].append(1)
            elif np.allclose(self.four, right1, atol=33) or np.allclose(self.fourr, right1, atol=33) == True:
                rlist[i].append(4)
            elif np.allclose(self.five, right1, atol=40) or np.allclose(self.fived, right1, atol=40) or np.allclose(self.fiver, right1, atol=40) == True:
                rlist[i].append(5)
            elif np.allclose(self.six, right1, atol=40) or np.allclose(self.sixr, right1, atol=40) == True:
                rlist[i].append(6)
            elif np.allclose(self.zero, right1, atol=33) or np.allclose(self.zeror, right1, atol=33) == True:
                rlist[i].append(0)
            elif np.allclose(self.niner, right1, atol=40) or np.allclose(self.nined, right1, atol=40) or np.allclose(self.ninerr, right1, atol=40) == True:
                rlist[i].append(9)
            elif np.allclose(self.threer, right1, atol=40) or np.allclose(self.threrr, right1, atol=40) == True:
                rlist[i].append(3)
            elif np.allclose(self.twor, right1, atol=40) or np.allclose(self.tworr, right1, atol=40) == True:
                rlist[i].append(2)
            elif np.allclose(self.eightr, right1, atol=33) or np.allclose(self.eightrrr, right1, atol=33) or np.allclose(self.eightd, right1, atol=33) or np.allclose(self.eightrr, right1, atol=33) == True:
                rlist[i].append(8)
            elif np.allclose(self.sevenr, right1, atol=33) == True:
                rlist[i].append(7)
            if np.allclose(self.one, right2, atol=40) or np.allclose(self.oner, right2, atol=40) == True:
                rlist[i].append(1)
            elif np.allclose(self.four, right2, atol=40) or np.allclose(self.fourr, right2, atol=33) == True:
                rlist[i].append(4)
            elif np.allclose(self.five, right2, atol=40) or np.allclose(self.fived, right2, atol=40) or np.allclose(self.fiver, right2, atol=40) == True:
                rlist[i].append(5)
            elif np.allclose(self.six, right2, atol=40) or np.allclose(self.sixr, right2, atol=40) == True:
                rlist[i].append(6)
            elif np.allclose(self.zero, right2, atol=27) or np.allclose(self.zeror, right2, atol=27) == True:
                rlist[i].append(0)
            elif np.allclose(self.niner, right2, atol=40) or np.allclose(self.nined, right2, atol=40) or np.allclose(self.ninerr, right2, atol=40) == True:
                rlist[i].append(9)
            elif np.allclose(self.threer, right2, atol=40) or np.allclose(self.threrr, right2, atol=40) == True:
                rlist[i].append(3)
            elif np.allclose(self.twor, right2, atol=40) or np.allclose(self.tworr, right2, atol=40) == True:
                rlist[i].append(2)
            elif np.allclose(self.eightr, right2, atol=40) or np.allclose(self.eightrrr, right2, atol=33) or np.allclose(self.eightd, right2, atol=40) or np.allclose(self.eightrr, right2, atol=40) == True:
                rlist[i].append(8)
            elif np.allclose(self.sevenr, right2, atol=40) == True:
                rlist[i].append(7)
            if np.allclose(self.one, right3, atol=47) or np.allclose(self.oner, right3, atol=47) == True:
                rlist[i].append(1)
            elif np.allclose(self.four, right3, atol=47) or np.allclose(self.fourr, right3, atol=40) == True:
                rlist[i].append(4)
            elif np.allclose(self.five, right3, atol=45) or np.allclose(self.fived, right3, atol=45) or np.allclose(self.fiver, right3, atol=45) == True:
                rlist[i].append(5)
            elif np.allclose(self.six, right3, atol=47) or np.allclose(self.sixr, right3, atol=40) == True:
                rlist[i].append(6)
            elif np.allclose(self.zero, right3, atol=47) or np.allclose(self.zeror, right3, atol=47) == True:
                rlist[i].append(0)
            elif np.allclose(self.niner, right3, atol=47) or np.allclose(self.nined, right3, atol=47) or np.allclose(self.ninerr, right3, atol=47) == True:
                rlist[i].append(9)
            elif np.allclose(self.threer, right3, atol=47) or np.allclose(self.threrr, right3, atol=40) == True:
                rlist[i].append(3)
            elif np.allclose(self.twor, right3, atol=47) or np.allclose(self.tworr, right3, atol=40) == True:
                rlist[i].append(2)
            elif np.allclose(self.eightr, right3, atol=47) or np.allclose(self.eightrrr, right3, atol=45) or np.allclose(self.eightd, right3, atol=47) or np.allclose(self.eightrr, right3, atol=47) == True:
                rlist[i].append(8)
            elif np.allclose(self.sevenr, right3, atol=47) == True:
                rlist[i].append(7)
            if np.allclose(self.one, right4, atol=47) or np.allclose(self.oner, right4, atol=47) == True:
                rlist[i].append(1)
            elif np.allclose(self.four, right4, atol=47) or np.allclose(self.fourr, right4, atol=40) == True:
                rlist[i].append(4)
            elif np.allclose(self.five, right4, atol=47) or np.allclose(self.fived, right4, atol=47) or np.allclose(self.fiver, right4, atol=47) == True:
                rlist[i].append(5)
            elif np.allclose(self.six, right4, atol=47)or np.allclose(self.sixr, right4, atol=40) == True:
                rlist[i].append(6)
            elif np.allclose(self.zero, right4, atol=47) or np.allclose(self.zeror, right4, atol=47) == True:
                rlist[i].append(0)
            elif np.allclose(self.niner, right4, atol=47) or np.allclose(self.nined, right4, atol=50) or np.allclose(self.ninerr, right4, atol=50) == True:
                rlist[i].append(9)
            elif np.allclose(self.threer, right4, atol=47) or np.allclose(self.threrr, right4, atol=40) == True:
                rlist[i].append(3)
            elif np.allclose(self.twor, right4, atol=47) or np.allclose(self.tworr, right4, atol=47) == True:
                rlist[i].append(2)
            elif np.allclose(self.eightr, right4, atol=47) or np.allclose(self.eightrrr, right4, atol=47) or np.allclose(self.eightd, right4, atol=47) or np.allclose(self.eightrr, right4, atol=47) == True:
                rlist[i].append(8)
            elif np.allclose(self.sevenr, right4, atol=47) == True:
                rlist[i].append(7)
        rlist.remove([])
        rlistik = []
        for i in rlist:
            magic = lambda nums: sum(digit * 10 ** (len(nums) - 1 - j)
            for j, digit in enumerate(i))
            rlistik.append(magic(i))
        # count = 0
        # kek = 0
        # for n in rlistik:
        #     if abs(2915 - n) > 30:
        #         count += 1
        # rlistik = [x for x in rlistik if abs(2915 - x) < 30]
        # print(count)
        # while kek < count:
        #     rlistik.append(int(np.mean(rlistik)))
        #     kek += 1
        # print('r', len(rlistik))
        return rlistik
    def toexcel(self):
        from openpyxl import Workbook
        wb = Workbook()
        dest_filename = 'Laba_BOLTSMAN.xlsx'
        ws = wb.create_sheet(title="Data")
        r = 1
        nya = 1
        en = 1
        for meow in self.touplist():
            ws.cell(row=r, column=1).value = meow
            r += 1
        for keks in self.todownlist():
            ws.cell(row=nya, column=2).value = keks
            nya += 1
        for statN in self.torightlist():
            ws.cell(row=en, column=3).value = statN
            en += 1
        wb.save(fi1lename=dest_filename)

a = Number()
# up1 = np.array(Image.open('v1/frame3.jpg'))[204:222, 113:126]
# a.toexcel()
print(a.touplist())
# print(a.todownlist())
# print(a.torightlist())
# plt.imshow(up1)
# plt.show()