# -*- coding: utf-8 -*-
"""
Created on Wed May 18 02:37:48 2016

@author: LENOVO
"""
from PIL import Image
import numpy as np
from statistics import mode
import matplotlib.pyplot as plt

i = Image.open('v1/frame0.jpg')
i1 = Image.open('v1/frame1699.jpg')
i2 = Image.open('v1/frame1566.jpg')
i3 = Image.open('v1/frame1555.jpg')
i4 = Image.open('v1/frame60.jpg')
iar = np.array(i)
qar = np.array(i1)
qar1 = np.array(i2)
qar2 = np.array(i3)
qar3 = np.array(i4)
qar4 = np.array(Image.open('v1/frame68.jpg'))
qar5 = np.array(Image.open('v1/frame22.jpg'))
qar6 = np.array(Image.open('v1/frame199.jpg'))
qar7 = np.array(Image.open('v1/frame719.jpg'))
qar8 = np.array(Image.open('v1/frame677.jpg'))
# print(iar[307][112])
# print(one)
keks = iar[204:222, 136:149]
one = iar[204:222, 136:149]
z = 1
while z < 21:
    z += 1
    one = np.vstack((one, np.array(Image.open('v1/frame%d.jpg' % z ))[204:222, 113:126]))
kek = 500
while kek < 530:
    kek += 1
    one = np.vstack((one, np.array(Image.open('v1/frame%d.jpg' % kek ))[204:222, 113:126]))
x = 1600
while x < 1630:
    x += 1
    one = np.vstack((one, np.array(Image.open('v1/frame%d.jpg' % x ))[204:222, 113:126]))
one = np.vsplit(one, 81)
two = np.array(Image.open('v1/frame1563.jpg'))[277:295, 136:149]
p = 1564
while p < 1584:
    p += 1
    two = np.vstack((two, np.array(Image.open('v1/frame%d.jpg' % p ))[277:295, 136:149]))
l = 0
while l < 20:
    l += 1
    two = np.vstack((two, np.array(Image.open('v2/frame%d.jpg' % l ))[278:296, 121:134]))
two = np.vsplit(two, 41)
three = np.array(Image.open('v2/frame27.jpg'))[278:296, 121:134]
s = 28
while s < 48:
    s += 1
    three = np.vstack((three, np.array(Image.open('v2/frame%d.jpg' % s ))[278:296, 121:134]))
n = 52
while n < 72:
    n += 1
    three = np.vstack((three, np.array(Image.open('v2/frame%d.jpg' % n ))[278:296, 121:134]))
three = np.vsplit(three, 41)
four = iar[278:296, 121:134]
m = 1
while m < 21:
    m += 1
    four = np.vstack((four, np.array(Image.open('v1/frame%d.jpg' % m ))[278:296, 121:134]))
q = 125
while q < 145:
    q += 1
    four = np.vstack((four, np.array(Image.open('v1/frame%d.jpg' % q ))[277:295, 136:149]))
four = np.vsplit(four, 41)
five = iar[204:222, 126:139]
w = 1
while w < 21:
    w += 1
    five = np.vstack((five, np.array(Image.open('v1/frame%d.jpg' % w ))[204:222, 126:139]))
y = 1
while y < 21:
    y += 1
    five = np.vstack((five, np.array(Image.open('v1/frame%d.jpg' % y ))[278:296, 107:120]))
b = 100
while b < 121:
    b += 1
    five = np.vstack((five, np.array(Image.open('v2/frame%d.jpg' % b ))[278:296, 107:120]))
u = 112
while u < 121:
    u += 1
    five = np.vstack((five, np.array(Image.open('v2/frame%d.jpg' % u ))[277:295, 136:149]))
five = np.vsplit(five, 71)
six = iar[204:222, 140:153]
j = 1
while j < 21:
    j += 1
    six = np.vstack((six, np.array(Image.open('v1/frame%d.jpg' % j ))[204:222, 140:153]))
g = 40
while g < 60:
    g += 1
    six = np.vstack((six, np.array(Image.open('v1/frame%d.jpg' % g ))[204:222, 140:153]))
v = 220
while v < 230:
    v += 1
    six = np.vstack((six, np.array(Image.open('v2/frame%d.jpg' % v ))[277:295, 136:149]))
six = np.vsplit(six, 51)
seven = np.array(Image.open('v2/frame15.jpg'))[277:295, 136:149]
v = 16
while v < 25:
    v += 1
    seven = np.vstack((seven, np.array(Image.open('v2/frame%d.jpg' % v ))[277:295, 136:149]))
seven = np.vsplit(seven, 10)
eight = np.array(Image.open('v2/frame200.jpg'))[204:222, 126:139]
v = 200
while v < 221:
    v += 1
    eight = np.vstack((eight, np.array(Image.open('v2/frame%d.jpg' % v ))[204:222, 126:139]))
eight = np.vsplit(eight, 22)
nine = iar[277:295, 136:149]
v = 1
while v < 21:
    v += 1
    nine = np.vstack((nine, np.array(Image.open('v1/frame%d.jpg' % v ))[277:295, 136:149]))
v = 243
while v < 253:
    v += 1
    nine = np.vstack((nine, np.array(Image.open('v2/frame%d.jpg' % v ))[277:295, 136:149]))
nine = np.vsplit(nine, 31)
zero = np.array(Image.open('v2/frame290.jpg'))[204:222, 113:126]
v = 291
while v < 321:
    v += 1
    zero = np.vstack((zero, np.array(Image.open('v2/frame%d.jpg' % v ))[204:222, 113:126]))
v = 1679
while v < 1699:
    v += 1
    zero = np.vstack((zero, np.array(Image.open('v1/frame%d.jpg' % v ))[277:295, 136:149]))
zero = np.vsplit(zero, 51)
rone = np.array(Image.open('v1/frame0.jpg'))[247:265, 404:417]
a = 1
while a < 19:
    a += 1
    rone = np.vstack((rone, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 404:417]))
a = 164
while a < 174:
    a += 1
    rone = np.vstack((rone, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
a = 1261
while a < 1271:
    a += 1
    rone = np.vstack((rone, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
a = 1321
while a < 1332:
    a += 1
    rone = np.vstack((rone, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 418:431]))
rone = np.vsplit(rone, 50)
rtwo = np.array(Image.open('v1/frame400.jpg'))[247:265, 376:389]
a = 401
while a < 421:
    a += 1
    rtwo = np.vstack((rtwo, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 376:389]))
a = 1297
while a < 1317:
    a += 1
    rtwo = np.vstack((rtwo, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
a = 500
while a < 511:
    a += 1
    rtwo = np.vstack((rtwo, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
rtwo = np.vsplit(rtwo, 52)
rthree = np.array(Image.open('v1/frame610.jpg'))[247:265, 404:417]
a = 611
while a < 631:
    a += 1
    rthree = np.vstack((rthree, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 404:417]))
a = 897
while a < 906:
    a += 1
    rthree = np.vstack((rthree, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
a = 1333
while a < 1343:
    a += 1
    rthree = np.vstack((rthree, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 418:431]))
a = 36
while a < 46:
    a += 1
    rthree = np.vstack((rthree, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
rthree = np.vsplit(rthree, 50)
rfour = np.array(Image.open('v1/frame610.jpg'))[247:265, 404:417]
a = 717
while a < 727:
    a += 1
    rfour = np.vstack((rfour, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
a = 700
while a < 720:
    a += 1
    rfour = np.vstack((rfour, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 376:389]))
a = 1249
while a < 1259:
    a += 1
    rfour = np.vstack((rfour, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
rfour = np.vsplit(rfour, 41)
rfive = np.array(Image.open('v2/frame400.jpg'))[247:265, 404:417]
a = 401
while a < 421:
    a += 1
    rfive = np.vstack((rfive, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 376:389]))
a = 0
while a < 7:
    a += 1
    rfive = np.vstack((rfive, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
a = 24
while a < 34:
    a += 1
    rfive = np.vstack((rfive, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
a = 60
while a < 70:
    a += 1
    rfive = np.vstack((rfive, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 418:431]))
rfive = np.vsplit(rfive, 48)
rsix = np.array(Image.open('v1/frame669.jpg'))[247:265, 418:431]
a = 669
while a < 679:
    a += 1
    rsix = np.vstack((rsix, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
a = 144
while a < 154:
    a += 1
    rsix = np.vstack((rsix, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
a = 180
while a < 190:
    a += 1
    rsix = np.vstack((rsix, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 418:431]))
rsix = np.vsplit(rsix, 31)
rseven = np.array(Image.open('v1/frame1524.jpg'))[247:265, 404:417]  
a = 1525
while a < 1545:
    a += 1
    rseven = np.vstack((rseven, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 404:417]))
a = 1521
while a < 1531:
    a += 1
    rseven = np.vstack((rseven, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
a = 120
while a < 140:
    a += 1
    rseven = np.vstack((rseven, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 390:403]))
a = 216
while a < 226:
    a += 1
    rseven = np.vstack((rseven, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
a = 396
while a < 406:
    a += 1
    rseven = np.vstack((rseven, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 418:431]))
rseven = np.vsplit(rseven, 71)
reight = np.array(Image.open('v1/frame1678.jpg'))[247:265, 390:403]
a = 1679
while a < 1699:
    a += 1
    reight = np.vstack((reight, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 390:403]))
a = 68
while a < 78:
    a += 1
    reight = np.vstack((reight, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 390:403]))
a = 811
while a < 821:
    a += 1
    reight = np.vstack((reight, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
a = 468
while a < 478:
    a += 1
    reight = np.vstack((reight, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 418:431]))
a = 485
while a < 505:
    a += 1
    reight = np.vstack((reight, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 390:403]))
a = 816
while a < 826:
    a += 1
    reight = np.vstack((reight, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
reight = np.vsplit(reight, 81)
rnine = np.array(Image.open('v1/frame1680.jpg'))[247:265, 404:417]
a = 1681
while a < 1691:
    a += 1
    rnine = np.vstack((rnine, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 404:417]))
a = 0
while a < 20:
    a += 1
    rnine = np.vstack((rnine, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 390:403]))
a = 500
while a < 510:
    a += 1
    rnine = np.vstack((rnine, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 390:403]))
a = 612
while a < 622:
    a += 1
    rnine = np.vstack((rnine, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 418:431]))
a = 700
while a < 720:
    a += 1
    rnine = np.vstack((rnine, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 390:403]))
a = 0
while a < 20:
    a += 1
    rnine = np.vstack((rnine, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 390:403]))
a = 0
while a < 10:
    a += 1
    rnine = np.vstack((rnine, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 404:417]))
rnine = np.vsplit(rnine, 101)
rzero = np.array(Image.open('v1/frame54.jpg'))[247:265, 404:417]
a = 55
while a < 65:
    a += 1
    rzero = np.vstack((rzero, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 404:417]))
a = 12
while a < 22:
    a += 1
    rzero = np.vstack((rzero, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 418:431]))
a = 80
while a < 90:
    a += 1
    rzero = np.vstack((rzero, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 418:431]))
a = 229
while a < 239:
    a += 1
    rzero = np.vstack((rzero, np.array(Image.open('v1/frame%d.jpg' % a ))[247:265, 404:417]))
a = 324
while a < 344:
    a += 1
    rzero = np.vstack((rzero, np.array(Image.open('v2/frame%d.jpg' % a ))[247:265, 390:403]))
rzero = np.vsplit(rzero, 61)
print('kek')
class Number(object):
    def __init__(self, vid):
        self.vid = vid
        self.one = one
        self.two = two
        self.three = three
        self.four = four
        self.five = five
        self.six = six
        self.seven = seven
        self.eight = eight
        self.nine = nine
        self.zero = zero
        self.rone = rone
        self.rtwo = rtwo
        self.rthree = rthree
        self.rfour = rfour
        self.rfive = rfive
        self.rsix = rsix
        self.rseven = rseven
        self.reight = reight
        self.rnine = rnine
        self.rzero = rzero

        #self.zero = qar[277:295, 136:149]  #0
        #self.zeror = qar3[247:265, 404:417] #0
        #self.zerod = qar[277:295, 136:149] #0
        #self.one = iar[204:222, 113:126]  #1
        #self.oner = iar[247:265, 404:417]  #1
        #self.oned = qar7[277:295, 136:149]  #1
        #self.twor = iar[247:265, 376:389]  #2
        #self.tworr = qar4[247:265, 418:431]  #2
        #self.twod = qar1[277:295, 136:149]  #2
        #self.threer = qar1[247:265, 418:431]  #3
        #self.threrr = qar6[247:265, 418:431]  #3
        #self.four = iar[278:296, 121:134]  #4
        #self.fourr = qar7[247:265, 418:431]  #4
        #self.five = iar[204:222, 126:139]  #5
        #self.fived = iar[278:296, 107:120]  #5
        #self.fivedd = qar7[278:296, 121:134]  #5
        #self.fiver = iar[247:265, 418:431]  #5
        #self.six = iar[204:222, 140:153]  #6
        #self.sixr = qar8[247:265, 418:431]  #6
        #self.sevenr = qar2[247:265, 418:431]  #7
        #self.eightr = qar[247:265, 390:403]  #8
        #self.eightrr = qar4[247:265, 390:403]
        #self.eightrrr = qar4[247:265, 418:431]  #8
        #self.eightd = qar3[277:295, 136:149] #8
        #self.niner = qar[247:265, 404:417]  #9
        #self.nined = iar[277:295, 136:149] #9
        #self.ninerr = iar[247:265, 390:403] #9
    def touplist(self):
        i = 0
        uplist = [[] for _ in range(1679)]
        while i < 1678:
            #up1 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[204:222, 113:126]
            up2 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[204:222, 126:139]
            up3 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[204:222, 140:153]
            i += 1
            #for elem in self.one:
            #    if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(1)
            #        break
            #for elem in self.two:
            #    if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(2)
            #        break
            #for elem in self.three:
            #    if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(3)
            #        break
            #for elem in self.four:
            #    if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(4)
            #        break
            #for elem in self.five:
            #    if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(5)
            #        break
            #for elem in self.six:
            #    if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(6)
            #        break
            #for elem in self.seven:
            #    if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(7)
            #        break
            #for elem in self.eight:
            #    if np.allclose(elem, up1, atol=10)  == True:
            #        uplist[i].append(8)
            #        break
            #for elem in self.nine:
            #    if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(9)
            #        break
            #for elem in self.zero:
           #     if np.allclose(elem, up1, atol=20)  == True:
            #        uplist[i].append(0)
            #        break
            for elem in self.one:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(1)
                    break
            for elem in self.two:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(2)
                    break
            for elem in self.three:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(3)
                    break
            for elem in self.four:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(4)
                    break
            for elem in self.five:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(5)
                    break
            for elem in self.six:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(6)
                    break
            for elem in self.seven:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(7)
                    break
            for elem in self.eight:
                if np.allclose(elem, up2, atol=25)  == True:
                    uplist[i].append(8)
                    break
            for elem in self.nine:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(9)
                    break
            for elem in self.zero:
                if np.allclose(elem, up2, atol=20)  == True:
                    uplist[i].append(0)
                    break
            for elem in self.one:
                if np.allclose(elem, up3, atol=20)  == True:
                    uplist[i].append(1)
                    break
            for elem in self.two:
                if np.allclose(elem, up3, atol=20)  == True:
                    uplist[i].append(2)
                    break
            for elem in self.three:
                if np.allclose(elem, up3, atol=20)  == True:
                    uplist[i].append(3)
                    break
            for elem in self.four:
                if np.allclose(elem, up3, atol=20)  == True:
                    uplist[i].append(4)
                    break
            for elem in self.five:
                if np.allclose(elem, up3, atol=20)  == True:
                    uplist[i].append(5)
                    break
            for elem in self.six:
                if np.allclose(elem, up3, atol=33)  == True:
                    uplist[i].append(6)
                    break
            for elem in self.seven:
                if np.allclose(elem, up3, atol=20)  == True:
                    uplist[i].append(7)
                    break
            for elem in self.eight:
                if np.allclose(elem, up3, atol=25)  == True:
                    uplist[i].append(8)
                    break
            for elem in self.nine:
                if np.allclose(elem, up3, atol=20)  == True:
                    uplist[i].append(9)
                    break
            for elem in self.zero:
                if np.allclose(elem, up3, atol=20)  == True:
                    uplist[i].append(0)
                    break             
        uplist.remove([])
        uplistik = []
        for i in uplist:
            magic = lambda nums: sum(digit * 10 ** (len(nums) - 1 - j)
                for j, digit in enumerate(i))
            uplistik.append(magic(i))
        count = 0
        kek = 0
        for n in uplistik:
            if abs(mode(uplistik) - n) > 30:
                count += 1
        uplistik = [x for x in uplistik if abs(mode(uplistik) - x) < 30]
        #print(count)
        while kek < count +1:
             uplistik.append(int(np.mean(uplistik)))
             kek += 1
        #print('u', len(uplistik))
        return uplistik
    def todownlist(self):
        i = 0
        downlist = [[] for _ in range(1679)]
        while i < 1678:
            down1 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[278:296, 107:120]
            down2 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[278:296, 121:134]
            down3 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[277:295, 136:149]
            i += 1
            for elem in self.one:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(1)
                    break
            for elem in self.two:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(2)
                    break
            for elem in self.three:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(3)
                    break
            for elem in self.four:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(4)
                    break
            for elem in self.five:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(5)
                    break
            for elem in self.six:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(6)
                    break
            for elem in self.seven:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(7)
                    break
            for elem in self.eight:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(8)
                    break
            for elem in self.nine:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(9)
                    break
            for elem in self.zero:
                if np.allclose(elem, down1, atol=30)  == True:
                    downlist[i].append(0)
                    break
            for elem in self.one:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(1)
                    break
            for elem in self.two:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(2)
                    break
            for elem in self.three:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(3)
                    break
            for elem in self.four:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(4)
                    break
            for elem in self.five:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(5)
                    break
            for elem in self.six:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(6)
                    break
            for elem in self.seven:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(7)
                    break
            for elem in self.eight:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(8)
                    break
            for elem in self.nine:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(9)
                    break
            for elem in self.zero:
                if np.allclose(elem, down2, atol=30)  == True:
                    downlist[i].append(0)
                    break
            for elem in self.one:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(1)
                    break
            for elem in self.two:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(2)
                    break
            for elem in self.three:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(3)
                    break
            for elem in self.four:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(4)
                    break
            for elem in self.five:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(5)
                    break
            for elem in self.six:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(6)
                    break
            for elem in self.seven:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(7)
                    break
            for elem in self.eight:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(8)
                    break
            for elem in self.nine:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(9)
                    break
            for elem in self.zero:
                if np.allclose(elem, down3, atol=30)  == True:
                    downlist[i].append(0)
                    break             
        downlist.remove([])
        downlistik = []
        for i in downlist:
            magic = lambda nums: sum(digit * 10 ** (len(nums) - 1 - j)
            for j, digit in enumerate(i))
            downlistik.append(magic(i))
        #print(np.mean(downlistik))
        count = 0
        kek = 0
        for n in downlistik:
            if abs(mode(downlistik) - n) > 30:
                count += 1
        downlistik = [x for x in downlistik if abs(532 - x) < 15]
        #print(count)
        while kek < count:
            downlistik.append(int(np.mean(downlistik)))
            kek += 1
        #print('d', len(downlistik))
        return downlistik
    def torightlist(self):
        i = 0
        rlist = [[] for _ in range(1679)]
        while i < 1678:
            right1 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[247:265, 376:389]
            right2 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[247:265, 390:403]
            right3 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[247:265, 404:417]
            right4 = np.array(Image.open('%s/frame%d.jpg' % (self.vid, i)))[247:265, 418:431]
            i += 1
            for elem in self.rone:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(1)
                    break
            for elem in self.rtwo:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(2)
                    break
            for elem in self.rthree:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(3)
                    break
            for elem in self.rfour:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(4)
                    break
            for elem in self.rfive:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(5)
                    break
            for elem in self.rsix:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(6)
                    break
            for elem in self.rseven:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(7)
                    break
            for elem in self.reight:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(8)
                    break
            for elem in self.rnine:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(9)
                    break
            for elem in self.rzero:
                if np.allclose(elem, right1, atol=39)  == True:
                    rlist[i].append(0)
                    break             
            for elem in self.rone:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(1)
                    break
            for elem in self.rtwo:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(2)
                    break
            for elem in self.rthree:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(3)
                    break
            for elem in self.rfour:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(4)
                    break
            for elem in self.rfive:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(5)
                    break
            for elem in self.rsix:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(6)
                    break
            for elem in self.rseven:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(7)
                    break
            for elem in self.reight:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(8)
                    break
            for elem in self.rnine:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(9)
                    break
            for elem in self.rzero:
                if np.allclose(elem, right2, atol=39)  == True:
                    rlist[i].append(0)
                    break             
            for elem in self.rone:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(1)
                    break
            for elem in self.rtwo:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(2)
                    break
            for elem in self.rthree:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(3)
                    break
            for elem in self.rfour:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(4)
                    break
            for elem in self.rfive:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(5)
                    break
            for elem in self.rsix:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(6)
                    break
            for elem in self.rseven:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(7)
                    break
            for elem in self.reight:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(8)
                    break
            for elem in self.rnine:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(9)
                    break
            for elem in self.rzero:
                if np.allclose(elem, right3, atol=39)  == True:
                    rlist[i].append(0)
                    break
            for elem in self.rone:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(1)
                    break
            for elem in self.rtwo:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(2)
                    break
            for elem in self.rthree:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(3)
                    break
            for elem in self.rfour:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(4)
                    break
            for elem in self.rfive:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(5)
                    break
            for elem in self.rsix:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(6)
                    break
            for elem in self.rseven:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(7)
                    break
            for elem in self.reight:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(8)
                    break
            for elem in self.rnine:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(9)
                    break
            for elem in self.rzero:
                if np.allclose(elem, right4, atol=39)  == True:
                    rlist[i].append(0)
                    break
                
        rlist.remove([])
        rlistik = []
        for i in rlist:
            magic = lambda nums: sum(digit * 10 ** (len(nums) - 1 - j)
            for j, digit in enumerate(i))
            rlistik.append(magic(i))
        count = 0
        kek = 0
        for n in rlistik:
            if abs(2915 - n) > 40:
                count += 1
        rlistik = [x for x in rlistik if abs(4950 - x) < 100]
        print(count)
        while kek < count:
            rlistik.append(int(np.mean(rlistik)))
            kek += 1
        print('r', len(rlistik))
        return rlistik
    def toexcel(self):
        from openpyxl import Workbook
        wb = Workbook()
        dest_filename = 'Laba_BOLTSMAN2.xlsx'
        ws = wb.create_sheet(title="Data")
        r = 1
        nya = 1
        en = 1
        for meow in self.touplist():
            ws.cell(row=r, column=1).value = meow
            r += 1
        for keks in self.todownlist():
            ws.cell(row=nya, column=2).value = keks
            nya += 1
        for statN in self.torightlist():
            ws.cell(row=en, column=3).value = statN
            en += 1
        wb.save(filename=dest_filename)

a = Number('v2')
b = Number('v1')
a.toexcel()
#print(b.todownlist())
print('done!')
#right1 = np.array(Image.open('v2/frame%d.jpg' % i))[247:265, 376:389]
# a.toexcel()
#print(a.touplist())
#print(a.todownlist())
#print(a.torightlist())
#plt.imshow(right1)
#plt.show()